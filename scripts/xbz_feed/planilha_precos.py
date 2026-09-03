"""Monta a planilha de precificacao do catalogo de clientes."""
import io
import json
import os

from openpyxl import Workbook
from openpyxl.comments import Comment
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

S = os.environ["SCRATCH"]
cat = json.load(io.open(os.path.join(S, "cat3.json"), encoding="utf-8"))
custos = json.load(io.open(os.path.join(S, "custos.json"), encoding="utf-8"))

FONTE = "Arial"
AZUL = Font(name=FONTE, size=10, color="0000FF")          # celula para preencher
PRETO = Font(name=FONTE, size=10)
CINZA = Font(name=FONTE, size=10, color="808080")
CAB = Font(name=FONTE, size=10, bold=True, color="FFFFFF")
TIT = Font(name=FONTE, size=13, bold=True, color="07253F")
NOTA = Font(name=FONTE, size=9, color="595959")
NOTA_B = Font(name=FONTE, size=9, bold=True, color="07253F")

FILL_CAB = PatternFill("solid", fgColor="07253F")
FILL_EDIT = PatternFill("solid", fgColor="FFF9E6")         # onde se digita
FILL_EXEMPLO = PatternFill("solid", fgColor="E8F5E9")
FILL_ALERTA = PatternFill("solid", fgColor="FDECEA")
BORDA = Border(*[Side(style="thin", color="D9D9D9")] * 4)

MOEDA = 'R$ #,##0.00'
MULT = '0.00"x"'

COLUNAS = [
    ("Código", 12),
    ("Descrição", 42),
    ("Situação", 11),
    ("Seção do catálogo", 26),
    ("Custo (R$)", 11),
    ("Qtd faixa 1", 10),
    ("Venda faixa 1 (R$)", 15),
    ("Qtd faixa 2", 10),
    ("Venda faixa 2 (R$)", 15),
    ("Qtd faixa 3", 10),
    ("Venda faixa 3 (R$)", 15),
    ("x1", 8),
    ("x2", 8),
    ("x3", 8),
]

LINHA_CAB = 10         # cabecalho da tabela (a legenda ocupa 5..8)
LINHA_EXEMPLO = 11     # linha de exemplo
LINHA_1 = 12           # primeiro produto


def escrever_aba(ws, linhas, titulo, subtitulo):
    ws.sheet_view.showGridLines = False

    ws["A1"] = titulo
    ws["A1"].font = TIT
    ws["A2"] = subtitulo
    ws["A2"].font = NOTA

    ws["A4"] = "Como preencher"
    ws["A4"].font = NOTA_B
    legenda = [
        "Edite apenas as colunas de fundo amarelo: Qtd faixa 1/2/3 e Venda faixa 1/2/3. "
        "O texto em azul é o valor que está no ar hoje - altere só o que quiser mudar.",
        "As faixas são a escada de quantidade do card: faixa 1 é o menor volume e faixa 3 é o "
        "melhor preço. Quase todo produto usa 20 / 50 / 100; caneta, sacola e o chaveiro 09824 "
        "usam 100 / 200 / 1000, porque o mínimo do fornecedor é 100.",
        "As colunas x1, x2 e x3 são calculadas (venda ÷ custo) e atualizam sozinhas conforme "
        "você digita. Servem para conferir a margem - não precisa preencher.",
        "Código, Descrição, Situação, Seção e Custo são só referência. Não altere o Código: "
        "é por ele que os preços voltam para o sistema.",
    ]
    for i, t in enumerate(legenda):
        ws.cell(row=5 + i, column=1, value="• " + t).font = NOTA
        ws.cell(row=5 + i, column=1).alignment = Alignment(vertical="center")

    # cabecalho
    for j, (nome, larg) in enumerate(COLUNAS, start=1):
        c = ws.cell(row=LINHA_CAB, column=j, value=nome)
        c.font = CAB
        c.fill = FILL_CAB
        c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        c.border = BORDA
        ws.column_dimensions[get_column_letter(j)].width = larg
    ws.row_dimensions[LINHA_CAB].height = 30

    # linha de exemplo
    exemplo = ["EXEMPLO", "Assim que a linha deve ficar preenchida", "—", "—",
               10.00, 20, 25.00, 50, 22.00, 100, 20.00]
    for j, v in enumerate(exemplo, start=1):
        c = ws.cell(row=LINHA_EXEMPLO, column=j, value=v)
        c.font = Font(name=FONTE, size=10, italic=True, color="1B5E20")
        c.fill = FILL_EXEMPLO
        c.border = BORDA
        if j in (5, 7, 9, 11):
            c.number_format = MOEDA
    for j, col in ((12, "G"), (13, "I"), (14, "K")):
        c = ws.cell(row=LINHA_EXEMPLO, column=j,
                    value=f'=IFERROR({col}{LINHA_EXEMPLO}/$E{LINHA_EXEMPLO},"")')
        c.font = Font(name=FONTE, size=10, italic=True, color="1B5E20")
        c.fill = FILL_EXEMPLO
        c.border = BORDA
        c.number_format = MULT

    for i, p in enumerate(linhas):
        r = LINHA_1 + i
        custo = custos.get(p["codigo"])
        valores = [
            p["codigo"],
            p["nome"],
            "No catálogo" if p["ativo"] else "Oculto",
            p["categoria_rotulo"] or p["categoria"],
            custo,
            p["faixa1_qtd"] or 20,
            float(p["faixa1_preco"]) if p["faixa1_preco"] is not None else None,
            p["faixa2_qtd"] or 50,
            float(p["faixa2_preco"]) if p["faixa2_preco"] is not None else None,
            p["faixa3_qtd"] or 100,
            float(p["faixa3_preco"]) if p["faixa3_preco"] is not None else None,
        ]
        for j, v in enumerate(valores, start=1):
            c = ws.cell(row=r, column=j, value=v)
            c.border = BORDA
            editavel = j in (6, 7, 8, 9, 10, 11)
            c.font = AZUL if editavel else (CINZA if j in (3, 4) else PRETO)
            if editavel:
                c.fill = FILL_EDIT
            if j in (5, 7, 9, 11):
                c.number_format = MOEDA
            if j == 1:
                c.alignment = Alignment(horizontal="left")

        # markup calculado, para a margem aparecer enquanto digita
        for j, col in ((12, "G"), (13, "I"), (14, "K")):
            c = ws.cell(row=r, column=j, value=f'=IFERROR({col}{r}/$E{r},"")')
            c.font = PRETO
            c.border = BORDA
            c.number_format = MULT

        # quando a faixa 1 nao esta em custo x 2,5, sinaliza: ou o fornecedor
        # mexeu no custo, ou o preco foi editado a mao
        if custo and p["faixa1_preco"] is not None:
            razao = float(p["faixa1_preco"]) / custo
            if abs(razao - 2.5) > 0.02:
                cel = ws.cell(row=r, column=12)
                cel.fill = FILL_ALERTA
                cel.comment = Comment(
                    f"Fora do padrão custo x 2,5 (está em {razao:.2f}x).\n"
                    f"Custo atual no fornecedor: R$ {custo:.2f}.\n"
                    "Ou o custo mudou desde que o catálogo foi montado, "
                    "ou o preço foi editado à mão.",
                    "Gift Web",
                )

    ws.freeze_panes = ws.cell(row=LINHA_1, column=3)
    ws.auto_filter.ref = f"A{LINHA_CAB}:N{LINHA_1 + len(linhas) - 1}"
    assert LINHA_CAB > 5 + len(legenda) - 1, "legenda invadiria o cabecalho"

    fim = LINHA_1 + len(linhas)
    ws.cell(row=fim + 1, column=1,
            value=f"{len(linhas)} produtos. Custo lido de products_cache "
                  "(preco_custo) em 04/09/2026; venda lida de catalogo_clientes.").font = NOTA


wb = Workbook()
ativos = [p for p in cat if p["ativo"]]
ocultos = [p for p in cat if not p["ativo"]]

ws1 = wb.active
ws1.title = "Catálogo"
escrever_aba(ws1, ativos, "Precificação — Catálogo Gift Web",
             f"{len(ativos)} produtos que o cliente vê hoje em giftwebbrindes.com.br/catalogo-clientes")

ws2 = wb.create_sheet("Ocultos")
escrever_aba(ws2, ocultos, "Precificação — produtos ocultos",
             f"{len(ocultos)} produtos que existem no cadastro mas estão desligados do catálogo. "
             "Preencha só se pretende reativá-los.")

destino = "data/precificacao_catalogo_giftweb.xlsx"
wb.save(destino)
print("gerado:", destino, "|", len(ativos), "ativos +", len(ocultos), "ocultos")
