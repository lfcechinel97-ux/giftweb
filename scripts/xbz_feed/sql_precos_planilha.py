"""Gera o SQL que aplica no catalogo os precos vindos da planilha preenchida.

Le a planilha de precificacao (a mesma gerada por planilha_precos.py, com as
faixas preenchidas a mao) e escreve um arquivo .sql para rodar no SQL editor do
Lovable Cloud - catalogo_clientes so aceita escrita de admin autenticado (RLS),
e este processo roda com a chave publica.

Sai um unico UPDATE ... FROM (VALUES ...) em vez de um UPDATE por produto: e
uma transacao so, entao ou todos os precos entram ou nenhum entra. Com 136
comandos soltos, uma queda no meio deixaria metade do catalogo com preco novo
e metade com o antigo.

Uso:
    python -m scripts.xbz_feed.sql_precos_planilha <planilha.xlsx>
"""
from __future__ import annotations

import sys
from decimal import Decimal
from pathlib import Path

from openpyxl import load_workbook

from . import config

LINHA_CAB, LINHA_1 = 10, 12
OBRIGATORIAS = (
    "Código", "Descrição", "Qtd faixa 1", "Venda faixa 1 (R$)",
    "Qtd faixa 2", "Venda faixa 2 (R$)", "Qtd faixa 3", "Venda faixa 3 (R$)",
)
SAIDA = config.ROOT / "supabase" / "migrations" / "20260904160000_catalogo_precos_planilha.sql"


def _cabecalho(ws) -> dict[str, int]:
    cab = {}
    for j in range(1, ws.max_column + 1):
        v = ws.cell(row=LINHA_CAB, column=j).value
        if v:
            cab[str(v).strip()] = j
    faltando = [c for c in OBRIGATORIAS if c not in cab]
    if faltando:
        raise SystemExit(f"{ws.title}: faltam as colunas {faltando}")
    return cab


def _dec(v) -> Decimal | None:
    if v is None or v == "":
        return None
    return Decimal(str(round(float(v), 2)))


def _int(v) -> int | None:
    if v is None or v == "":
        return None
    return int(round(float(v)))


def ler(caminho: Path) -> list[dict]:
    wb = load_workbook(caminho, data_only=True)
    linhas: list[dict] = []
    for ws in wb.worksheets:
        cab = _cabecalho(ws)
        for r in range(LINHA_1, ws.max_row + 1):
            cod = ws.cell(row=r, column=cab["Código"]).value
            # a nota de rodape tambem mora na coluna A; produto tem descricao
            if not cod or not ws.cell(row=r, column=cab["Descrição"]).value:
                continue
            cod = str(cod).strip()
            if cod.upper() == "EXEMPLO":
                continue
            linhas.append({
                "codigo": cod, "aba": ws.title, "linha": r,
                "q1": _int(ws.cell(row=r, column=cab["Qtd faixa 1"]).value),
                "p1": _dec(ws.cell(row=r, column=cab["Venda faixa 1 (R$)"]).value),
                "q2": _int(ws.cell(row=r, column=cab["Qtd faixa 2"]).value),
                "p2": _dec(ws.cell(row=r, column=cab["Venda faixa 2 (R$)"]).value),
                "q3": _int(ws.cell(row=r, column=cab["Qtd faixa 3"]).value),
                "p3": _dec(ws.cell(row=r, column=cab["Venda faixa 3 (R$)"]).value),
            })
    return linhas


def validar(linhas: list[dict]) -> list[str]:
    """Erro de digitacao aqui vira preco errado no ar, entao nada passa sem conferencia."""
    erros: list[str] = []
    vistos: set[str] = set()
    for l in linhas:
        onde = f"{l['aba']} linha {l['linha']} ({l['codigo']})"
        if l["codigo"] in vistos:
            erros.append(f"{onde}: codigo repetido")
        vistos.add(l["codigo"])
        qs = [l["q1"], l["q2"], l["q3"]]
        ps = [l["p1"], l["p2"], l["p3"]]
        if any(v is None for v in qs + ps):
            erros.append(f"{onde}: campo de faixa em branco")
            continue
        if not (qs[0] < qs[1] < qs[2]):
            erros.append(f"{onde}: quantidades {qs} nao sao crescentes")
        if not (ps[0] >= ps[1] >= ps[2]):
            erros.append(f"{onde}: precos {ps} nao caem conforme o volume sobe")
        if any(p <= 0 for p in ps):
            erros.append(f"{onde}: preco zero ou negativo")
    return erros


def main() -> int:
    if len(sys.argv) < 2:
        raise SystemExit("uso: python -m scripts.xbz_feed.sql_precos_planilha <planilha.xlsx>")
    caminho = Path(sys.argv[1])
    linhas = ler(caminho)
    print(f"[PRECOS] {len(linhas)} produtos lidos de {caminho.name}")

    erros = validar(linhas)
    if erros:
        print(f"[PRECOS] {len(erros)} problemas - nada foi gerado:")
        for e in erros[:30]:
            print("   -", e)
        return 1

    # sempre com duas casas: alem de ficar legivel na revisao, deixa a coluna
    # inteira como numeric no VALUES, sem depender de inferencia de tipo
    valores = ",\n".join(
        f"    ('{l['codigo']}', {l['q1']}, {l['p1']:.2f}, "
        f"{l['q2']}, {l['p2']:.2f}, {l['q3']}, {l['p3']:.2f})"
        for l in linhas
    )
    sql = f"""-- Precos do catalogo de clientes vindos da planilha preenchida.
--
-- Fonte: {caminho.name}, {len(linhas)} produtos (catalogo + ocultos).
-- Gerado por scripts/xbz_feed/sql_precos_planilha.py - nao editar a mao.
--
-- Um UPDATE so, e nao um por produto: e uma transacao unica, entao ou todos os
-- precos entram ou nenhum entra. Com 136 comandos soltos, uma falha no meio
-- deixaria metade do catalogo com preco novo e metade com o antigo.
--
-- "preco" e atualizado junto com a faixa 1. Ele nao aparece para o cliente
-- enquanto as tres faixas estiverem preenchidas (e a reserva de quando faltar
-- alguma), mas e o campo "Preco exibido" do /admin - deixa-lo com o valor
-- antigo faria o painel mostrar um numero que nao existe mais.
--
-- As mesmas linhas alimentam o catalogo do cliente e o /admin: as duas telas
-- leem catalogo_clientes, entao nao existe segundo lugar para atualizar.

UPDATE public.catalogo_clientes AS c
SET faixa1_qtd   = v.q1,
    faixa1_preco = v.p1,
    faixa2_qtd   = v.q2,
    faixa2_preco = v.p2,
    faixa3_qtd   = v.q3,
    faixa3_preco = v.p3,
    preco        = v.p1
FROM (VALUES
{valores}
) AS v(codigo, q1, p1, q2, p2, q3, p3)
WHERE c.codigo = v.codigo;

-- Conferencia: deve voltar {len(linhas)} atualizados e 0 fora do padrao.
SELECT count(*) FILTER (WHERE faixa1_preco IS NOT NULL)                AS com_faixa,
       count(*) FILTER (WHERE faixa1_qtd >= faixa2_qtd
                           OR faixa2_qtd >= faixa3_qtd)                AS qtd_fora_de_ordem,
       count(*) FILTER (WHERE faixa1_preco < faixa2_preco
                           OR faixa2_preco < faixa3_preco)             AS preco_fora_de_ordem
FROM public.catalogo_clientes;
"""
    SAIDA.write_text(sql, encoding="utf-8")
    print(f"[PRECOS] SQL em {SAIDA}")
    return 0


if __name__ == "__main__":
    sys.exit(main())
